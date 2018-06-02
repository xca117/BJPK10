#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Date    : 2018-05-31 08:33:35
# @Author  : xca117 (408114416@qq.com)
# @Link    : *
# @Version : $Id$

import os
import requests
import time
from bs4 import BeautifulSoup
from openpyxl import *
import setup as gui

headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
'Accept-Encoding': 'gzip, deflate',
'Accept-Language': 'zh-CN,zh;q=0.9',
'Cache-Control': 'max-age=0',
'Connection': 'keep-alive',
'Host': 'www.bwlc.net',
'Referer': 'http://www.bwlc.net/',
'Upgrade-Insecure-Requests': '1',
'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36'
}
datas = []
if os.path.exists(r'北京PK10数据.xlsx'):
    workbook = load_workbook(r'北京PK10数据.xlsx')
    sh = workbook['Data']
    if sh.max_row >= 361:
        for row in sh.rows:
            datas.append([cell.value for cell in row])
    workbook.remove(sh)
else:
    workbook = Workbook()
sh = workbook.create_sheet('Data',index=0)
sh.append(['时间','期号','一','二','三','四','五',
        '六','七','八','九','十'])

start_url = 'http://www.bwlc.net/bulletin/trax.html?page={page}'
now_row = 0

def request(url):
    r = requests.get(url,headers=headers)
    return r.text
def parse(url):
    res = request(url)
    soup = BeautifulSoup(res,'lxml')
    result_1 = soup.find_all('div',class_='lott_cont')
    items = []
    for result in result_1:
        result_2 = result.find_all('tr')
        for i in range(1,len(result_2)):
            result = result_2[i].find_all('td')
            items.append((result[2].string,result[0].string,result[1].string,))
    # print(items)
    return items

def save(items):
    global now_row
    con = 0
    if items:
        # 第一行为标题,数据在第二行 datas[1],开奖号从第3列值开始
        for item in items:
            # 如果采集的数据与已有的第一行期号及开奖号数据相同,则停止采集
            if datas:
                if item[1] == datas[1][1] and item[2] == ','.join(datas[1][2:]):
                    # 358 减 当前已采集行 = t
                    t = 358 - now_row
                    for i in range(1,t+1):
                        # 开始保存datas[i]
                        # print(datas[i])
                        sh.append(datas[i])
                    con = 1
                    break
            num = item[2].split(',')
            sh.append([item[0],item[1],num[0],num[1],num[2],num[3],num[4],num[5],num[6],num[7],num[8],num[9]])
            now_row += 1
        workbook.save(r'北京PK10数据.xlsx')
    return True if con == 0 else False

def find_now_data():
    item = ''
    items = parse(start_url.format(page=1))
    if items:
        item = items[0][2]
        return item
    return items


def start_all():
    for i in range(1,13):
        items = parse(start_url.format(page=i))
        if not save(items):
            break
        time.sleep(1)
        # print(i)

# find_now_data()
# print(parse(start_url.format(page=1)))